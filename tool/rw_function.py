import openpyxl
import numpy as np
import os
import matplotlib.pyplot as plt
import ast
import math
import tool.result_function as rf
import tool.robot_function as rof


def IFX_SET(F,X,object_items):
    iNo =np.array([[i] for i in range(X.shape[0])])
    IF = np.hstack((iNo, F))
    IFX = np.hstack((IF, X))
    x_name = []
    for i in range(X.shape[1]):
        x_name.append("x{}".format(i+1))
    header_name = ["iNo"] + object_items + x_name
    return IFX,header_name

def confirm_file(file_name):
    is_file = os.path.isfile(file_name)
    if is_file==False:
        wb = openpyxl.Workbook()
        wb.save(file_name)
        
def compare_sheets(current_dicts,exist_dicts):
    for c,e in zip(current_dicts,exist_dicts):
        value = current_dicts[c]
        if type(value) == np.ndarray:
            value = value.tolist()
        if type(value) == list:
            value = str(value)
        if (value==exist_dicts[e])==False:
            return None
    return True

def judge_gaparas(wb,current_dicts,ga_para_keys):
    sheets = wb.sheetnames
    for sh in sheets:
        ws = wb[sh]
        for row in ws.iter_rows(min_row=2, max_row=2,max_col=len(ga_para_keys)):
            data=[]
            for cell in row:
                data.append(cell.value)
        exist_dicts = dict(zip(ga_para_keys, data))
        sheet_name = compare_sheets(current_dicts,exist_dicts)
        if sheet_name!=None:
            return sh,None
    return None,len(sheets)
        
def save_gaparas(file_name,ga_para_keys,ga_para_values):
    wb = openpyxl.load_workbook(file_name)
    current_dicts = dict(zip(ga_para_keys,ga_para_values))
    sheet_name,sheet_num = judge_gaparas(wb,current_dicts,ga_para_keys)
    if sheet_name == None:
        wb.create_sheet('case{}'.format(sheet_num-1))
        sheets = wb.sheetnames
        ws = wb[sheets[-1]]
        for i in range(len(ga_para_keys)):
            ws.cell(row=1, column=i+1).value = ga_para_keys[i]
            value = ga_para_values[i]
            if type(value) == np.ndarray:
                value = value.tolist()
            if type(value) == list:
                value = str(value)
            ws.cell(row=2, column=i+1).value = value
        wb.save(file_name)
        return sheets[-1]
    return sheet_name

def save_datas(file_name,F,X,object_items,sheet_name,seed_no):
    wb = openpyxl.load_workbook(file_name)
    ws = wb[sheet_name]
    IFX,header_name = IFX_SET(F,X,object_items)
    max_row = ws.max_row
    ws.cell(row=max_row+1, column=1).value = "seed{}".format(seed_no)
    max_row = ws.max_row
    for i in range(len(header_name)):
        ws.cell(row=max_row+1, column=i+1).value = header_name[i]
    max_row = ws.max_row
    for i in range(IFX.shape[0]):
        for j in range(IFX.shape[1]):
            ws.cell(row=max_row+1+i, column=j+1).value = IFX[i][j]
    wb.save(file_name)

def evalues_course_datas(file_name,sheet_name,evalues,course_all):
    wb = openpyxl.load_workbook(file_name)
    ws = wb[sheet_name]
    max_row = ws.max_row
    for i in range(len(evalues)):
        ws.cell(row=max_row+1, column=i+1).value = evalues[i]
    max_row = ws.max_row
    for i in range(course_all.shape[0]):
        for j in range(course_all.shape[1]):
            ws.cell(row=max_row+1+i, column=j+1).value = course_all[i][j]
    max_row = ws.max_row
    tmp = course_all.mean(axis=0)
    for k in range(len(tmp)):
        ws.cell(row=max_row+1, column=k+1).value = tmp[k]
    wb.save(file_name)

def get_indivi_arr(file_name,sheet_name,size_obj):
    wb = openpyxl.load_workbook(file_name)
    ws = wb[sheet_name]
    datas = []
    head_datas = []
    for row in ws.iter_rows(max_row=2):
        data=[]
        for cell in row:
            data.append(cell.value)
        head_datas.append(data)

    for row in ws.iter_rows(min_row=3):
        data=[]
        for cell in row:
            data.append(cell.value)
        datas.append(data)
        
    indivi_list = []
    compre_list = []
    count = 0
    count2 = 0
    compre = 0
    for d in datas:
        if type(d[0])==int or type(d[0])==float:
            d1_tmp = d[0]
            d[0]=count
            indivi_list.append(d)
            count += 1
            if compre == 1:
                d[0]=d1_tmp
                count2 += 1
                compre_list.append(d)
        elif d[0]=="comprehensive_start":
            compre = 1
        elif d[0]=="comprehensive_end":
            compre_arr = np.array(compre_list)
            compre_arr = compre_arr[:, np.all(compre_arr!=None, axis=0)]
            return compre_arr,head_datas
    
    ids_arr = rank_dicision(np.array(indivi_list),size_obj)
    ids_new = ids_arr[np.argsort(ids_arr[:, -1])]

    max_row = ws.max_row
    ws.cell(row=max_row+2, column=1).value = "comprehensive_start"
    max_row = ws.max_row
    for i in range(ids_new.shape[0]):
        for j in range(ids_new.shape[1]):
            ws.cell(row=max_row+1+i, column=j+1).value = ids_new[i][j]
    max_row = ws.max_row
    ws.cell(row=max_row+1, column=1).value = "comprehensive_end"
    wb.save(file_name)
    return  ids_new,head_datas

def rank_dicision(indivi_obj,size_obj):
    ids = []
    for id in indivi_obj:
        rank = 1
        for i in range(len(indivi_obj)):
            if id[0]==i:
                pass
            else:
                if np.all(id[1:size_obj+1] > indivi_obj[i][1:size_obj+1]):
                    rank += 1
        id = id[~(id == None)]
        tmp = id.tolist()
        tmp.append(rank)
        ids.append(tmp)
    ids_arr = np.array(ids)
    return ids_arr

def option_arr(file_name,sheet_name,option):
    wb = openpyxl.load_workbook(file_name)
    ws = wb[sheet_name]
    datas = []
    head_datas = []
    for row in ws.iter_rows(max_row=2):
        data=[]
        for cell in row:
            data.append(cell.value)
        head_datas.append(data)
    size_obj = len(head_datas[1][16].split("+"))
    for row in ws.iter_rows(min_row=3):
        data=[]
        for cell in row:
            data.append(cell.value)
        datas.append(data)
    compre_list = []
    compre = 0
    for d in datas:
        if type(d[0])==int or type(d[0])==float:
            if compre == 1:
                compre_list.append(d)
        elif d[0]==option+"_start":
            compre = 1
        elif d[0]==option+"_end":
            break
    return  np.array(compre_list),head_datas,size_obj

def train_course_evalues(file_name,num):
    case_set = []
    for i in range(num):
        sheet_name = "case{}".format(i)
        print(sheet_name)
        indivi_arr,head_datas,size_obj = option_arr(file_name,sheet_name,"comprehensive")
        indivi_arr = delete_none(indivi_arr)
        only_one_arr = np.delete(indivi_arr,np.where(indivi_arr[:,-1]!=1),axis=0)
        case_set.append([only_one_arr,head_datas[1][16].split("+")])
    return case_set

def print_compare_set(case_set,items,evalues,num):
    compare_sets = []
    num = 0
    for i in range(num):
        count = 0
        ind_case = case_set[i][1]
        index_set =[]
        for it in [evalues[i] for i in items]:
            if it in ind_case:
                index_set.append(ind_case.index(it)+1)
                count += 1
        if count == len(items):
            compare_sets.append([i,index_set,ind_case])
            print(num,i,ind_case)
            num += 1
    return compare_sets
            
def train_compare_display(case_set,compare_sets,sets,display):
    set1 = sets[0]
    set2 = sets[1]
    set_no1 = compare_sets[set1][0]
    set_no2 = compare_sets[set2][0]
    x1 = compare_sets[set1][1][0]
    y1 = compare_sets[set1][1][1]
    x2 = compare_sets[set2][1][0]
    y2 = compare_sets[set2][1][1]
    if display:
        plt.scatter(case_set[set_no1][0][:,x1],case_set[set_no1][0][:,y1],color="blue")
        plt.scatter(case_set[set_no2][0][:,x2],case_set[set_no2][0][:,y2],color="red")
        plt.show()

def output_divi(rule,set_types,Kch,out_level):
    division_list = []
    if set_types[-1]==0:
        end_shape = out_level
    else:
        end_shape = out_level
    shape_list = list(rule.shape)+[end_shape]
    if Kch==False:
        shape_list = [2]*len(shape_list)    
    for r,t in zip(shape_list,set_types):
        if t==0:
            r = r//2+1
        division_list.append(r-1)
    return division_list

def Ks_divi(pli,divi_li):
    Ks_list = []
    n = 0
    for d in divi_li:
        Ks_list.append(pli[n:n+d:1])
        n += d
    return Ks_list

####直線コース
def straight_course(robot,Kis):
    vdatas = []
    lms = np.array([[0,0],[0,0.5],[0,1],[0,1.5],[0,2]])
    for i in range(3):
        for j in range(8):
            fpos = [0.5*i,0,j*math.pi/4]
            rf.save_drive_orbit(robot,Kis,fpos,lms,20,False,3)
            t = robot.arrvial_time
            v = save_control_values(robot)
            v.insert(0,t)
            vdatas.append(v)
    return np.array(vdatas)

###曲線コース
def curve_course(robot,Kis):
    vdatas = []
    for j in range(3):
        r = 0.2+0.4*j
        lms_list = []
        for i in range(0,18):
            theta = i*math.pi/8
            x = r*math.cos(theta)
            y = r*math.sin(theta)
            lms_list.append([x,y])
        lms = np.array(lms_list)
        for k in range(8):
            fpos = [r,0,k*math.pi/4]
            rf.save_drive_orbit(robot,Kis,fpos,lms,50,False,16)
            t = robot.arrvial_time
            v = save_control_values(robot)
            v.insert(0,t)
            vdatas.append(v)
    return np.array(vdatas)

#制御評価
def save_control_values(obj_control):
    x = obj_control.li1
    y = obj_control.li2
    t = obj_control.li3
    n = obj_control.now_nos
    v = obj_control.liv
    nd = obj_control.li_direction
    dis = obj_control.error_distance_calc([1,x,y,t,n])
    di_delta = obj_control.direction_shake_calc(t)
    speed_avg =obj_control.speed_average_calc(v)
    next_di = obj_control.next_direction_error_calc(nd)
    return [dis,di_delta,speed_avg,next_di]

def none_judge_mean(value_arr):
    if np.any(value_arr==None):
        return np.full((5,),None)
    else:
        return value_arr.mean(axis=0)

#個体ごとに評価測定
def ind_evalue(ind,head_data,size_obj,divi_list):
    robot1 = rof.obj_func(fuzzy_rule1=np.array(head_data["fuzzy_rule1"]),
             set_types1 = head_data["set_types1"],
             out_level1 = head_data["out_level1"],
             fuzzy_rule2 = np.array(head_data["fuzzy_rule2"]),
             set_types2 = head_data["set_types2"],
             out_level2 = head_data["out_level2"],
             object_items = head_data["object_items"].split("+"))
    Kis = Ks_divi(ind[size_obj+1:],divi_list)
    
    vdatas_st_arr = straight_course(robot1,Kis)
    vdatas_cu_arr = curve_course(robot1,Kis)
    course1 = none_judge_mean(vdatas_st_arr[:][0:8])
    course2 = none_judge_mean(vdatas_st_arr[:][9:16])
    course3 = none_judge_mean(vdatas_st_arr[:][17:24])
    course4 = none_judge_mean(vdatas_cu_arr[:][0:8])
    course5 = none_judge_mean(vdatas_cu_arr[:][9:16])
    course6 = none_judge_mean(vdatas_cu_arr[:][17:24])
    course_all = np.array([course1,course2,course3,course4,course5,course6])
    all_avg = none_judge_mean(course_all)
    save_datas = np.r_[all_avg,course1,course2,course3,course4,course5,course6]
    ind_save_datas = np.insert(ind,1,save_datas)
    return ind_save_datas

def course_evalue(file_name,sheet_name):
    indivi_arr,head_datas,size_obj = option_arr(file_name,sheet_name,"comprehensive")
    indivi_arr = delete_none(indivi_arr)
    indivi_arr = np.delete(indivi_arr,np.where(indivi_arr[:,-1]!=1),axis=0)

    
    for i in [5,6,7,8,11,12,13,14]:
        s = head_datas[1][i]
        head_datas[1][i]= ast.literal_eval(s)
    head_data = dict(zip(head_datas[0],head_datas[1]))
    divi1 =  output_divi(np.array(head_data["fuzzy_rule1"]),head_data["set_types1"],head_data["Kch1"],head_data["out_level1"])
    divi2 =  output_divi(np.array(head_data["fuzzy_rule2"]),head_data["set_types2"],head_data["Kch2"],head_data["out_level2"])
    divi_list = divi1+divi2
    
    all_datas = []
    for i in range(len(indivi_arr)):
        ind = indivi_arr[i]
        all_datas.append(ind_evalue(ind,head_data,size_obj,divi_list))
        
    v_arr = np.array(all_datas)
    v_arr = np.delete(v_arr, [36+i for i in range(size_obj)], 1)
    v_arr = np.delete(v_arr, -1, 1)
    return v_arr

def evalues_course_datas(file_name,sheet_name,evalues,save_datas):
    wb = openpyxl.load_workbook(file_name)
    ws = wb[sheet_name]
    max_row = ws.max_row
    ws.cell(row=max_row+1, column=1).value = "evalue_course_start"
    max_row = ws.max_row
    ws.cell(row=max_row+1, column=1).value = "iNo"
    for j in range(7):
        for i in range(len(evalues)):
            ws.cell(row=max_row+1, column=5*j+i+2).value = evalues[i]+str(j)
    max_row = ws.max_row
    
    for i in range(save_datas.shape[0]):
        for j in range(save_datas.shape[1]):
            ws.cell(row=max_row+1+i, column=j+1).value = save_datas[i][j]
    max_row = ws.max_row
    ws.cell(row=max_row+1, column=1).value = "evalue_course_end"
    wb.save(file_name)
    
def delete_none(indivi_arr):
    for i in range(len(indivi_arr[0])):
        if indivi_arr[0][i]==None:
            return indivi_arr[:,0:i]
    if i==len(indivi_arr[0])-1:
        return indivi_arr[:,0:]
    
def save_test_course_evalue(file_name,num):
    sheet_name = "case"
    evalues = ["end_time", "err_distance", "direction_shake","speed","next_direction"]
    #評価コースで制御を評価
    for i in range(0,num):
        sname = "case{}".format(i)
        tmp = course_evalue(file_name,sname)
        #データ保存
        evalues_course_datas(file_name,sname,evalues,tmp)
        print("No."+str(i))

def test_course_evalues(file_name,num):
    case_set = []
    sheet_name = "case"
    wb = openpyxl.load_workbook(file_name)
    for i in range(0,num):
        sname = sheet_name+str(i)
        ws = wb[sname]
        head_datas = []
        for row in ws.iter_rows(max_row=2):
            data=[]
            for cell in row:
                data.append(cell.value)
                head_datas.append(data)
        indivi_arr,head_datas,size_obj = option_arr(file_name,sname,"evalue_course")
        datas = np.delete(indivi_arr,np.where(indivi_arr[:,1]==None),axis=0) #Noneある個体は消す
        case_set.append([datas,head_datas[1][16].split("+")])
        print("No."+str(i))
    return case_set

def test_compare_display(case_set,sets,items_no,display,dim1):
    if dim1:
        for j in items_no:
            for i in sets:
                xname = "case{}".format(i)
                x = [xname for i in range(len(case_set[i][0]))]
                y = case_set[i][0][:,j+1]
                plt.scatter(x,y)
            plt.show()
            
    if display:
        plt.scatter(case_set[sets[0]][0][:,items_no[0]+1],case_set[sets[0]][0][:,items_no[1]+1],color="blue")
        plt.scatter(case_set[sets[1]][0][:,items_no[0]+1],case_set[sets[1]][0][:,items_no[1]+1],color="red")
        plt.show()
